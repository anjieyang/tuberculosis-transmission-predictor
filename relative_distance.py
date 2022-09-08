from common_modules import read_data_from_xls, Building_Number
from openpyxl import Workbook, load_workbook
import xlrd


def xls_to_xlsx(xls_path):
    """
    This function takes the path of a .xls file and save the same-content .xlsx file into
    the same directory.
    :param xls_path: The path of original .xls file.
    """
    xlsx_path = xls_path.split(".")[0] + ".xlsx"
    book_xls = xlrd.open_workbook(xls_path)
    book_xlsx = Workbook()

    sheet_names = book_xls.sheet_names()
    for sheet_index, sheet_name in enumerate(sheet_names):
        sheet_xls = book_xls.sheet_by_name(sheet_name)
        if sheet_index == 0:
            sheet_xlsx = book_xlsx.active
            sheet_xlsx.title = sheet_name
        else:
            sheet_xlsx = book_xlsx.create_sheet(title=sheet_name)

        for row in range(0, sheet_xls.nrows):
            for col in range(0, sheet_xls.ncols):
                sheet_xlsx.cell(row=row + 1, column=col + 1).value = sheet_xls.cell_value(row, col)

        book_xlsx.save(xlsx_path)


PATH = "geo coordinates/"
FILE = "Arviat"
if __name__ == "__main__":
    try:
        book = load_workbook(PATH + FILE + ".xlsx")
    except FileNotFoundError:
        xls_to_xlsx(PATH + FILE + ".xls")
        book = load_workbook(PATH + FILE + ".xlsx")

    buildings = read_data_from_xls(PATH + FILE + ".xls")

    try:
        sheet = book["relative_distances"]
    except KeyError:
        book.create_sheet("relative_distances")
        sheet = book["relative_distances"]

    sheet.cell(1, 1, "Building ID")
    for i in range(0, len(buildings)):
        # print("Row: {} Building ID: {}".format(i + 1, buildings[i].get_id()))
        # Add each building id to rows
        sheet.cell(i + 2, 1, buildings[i].get_id())
        # Add each building id to columns
        sheet.cell(1, i + 2, buildings[i].get_id())

    # Calculate the process time
    process_times = 0
    for i in range(1, len(buildings) + 1):
        process_times += i

    processed = 0
    for i in range(len(buildings)):
        for j in range(len(buildings)):
            # Avoid repeated calculation
            if i > j:
                distance = sheet.cell(j + 2, i + 2).value
                sheet.cell(i + 2, j + 2, distance)
                continue
            processed += 1
            sheet.cell(i + 2, j + 2, Building_Number.find_distance(buildings[i], buildings[j]))
            print("Processing Progress: {} / {} --> {}%\n".format(processed, process_times,
                                                                  int(processed / process_times * 100)))

        # Avoid losing all processed data
        book.save(PATH + FILE + ".xlsx")
